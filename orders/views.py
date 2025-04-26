#orders/views.py
from django.shortcuts import render, redirect, HttpResponse
from .models import Order, OrderItem
from carts.models import Cart, CartItem

from .forms import OrderForm
from .export_to_gsheets import export_orders_to_gsheets
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.pagesizes import A3
from django.shortcuts import get_list_or_404
from carts.views import clear_cart

def checkout(request):
    orders = Order.objects.all()
    return render(request, 'orders.html', {'orders': orders})

def order_complete(request):
    return render(request, 'order_complete.html')


def export_orders_pdf(request):
    queryset = get_list_or_404(Order)
    return export_orders_view(request, queryset, 'pdf')

def export_orders_xlsx(request):
    queryset = get_list_or_404(Order)
    return export_orders_view(request, queryset, 'xlsx')

def export_orders(request):
    queryset = get_list_or_404(Order)
    export_orders_view(request, queryset, 'pdf')
    return export_orders_view(request, queryset, 'xlsx')

def export_orders_view(request, queryset, format):
    orders = queryset

    if format == 'xlsx':
        # Your existing Excel code...
        wb = Workbook()
        ws = wb.active

        headers = ['First Name', 'Last Name', 'Phone', 'Email', 'Date', 'Address', 'Status']
        for col_num, header in enumerate(headers, 1):
            col_letter = ws.cell(row=1, column=col_num).column_letter
            ws['{}1'.format(col_letter)] = header
            ws.column_dimensions[col_letter].width = 15

        for row_num, order in enumerate(orders, 2):
            ws.cell(row=row_num, column=1, value=order.first_name)
            ws.cell(row=row_num, column=2, value=order.last_name)
            ws.cell(row=row_num, column=3, value=order.order_phone)
            ws.cell(row=row_num, column=4, value=order.order_username)
            ws.cell(row=row_num, column=5, value=order.order_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=7, value=order.order_address)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=orders.xlsx'
        wb.save(response)

    elif format == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=orders.pdf'

        doc = SimpleDocTemplate(response, pagesize=landscape(A3))

        data = []
        headers = ['First Name', 'Last Name', 'Phone', 'Email', 'Date', 'Address', 'Status']
        data.append(headers)

        for order in orders:
            row = [order.first_name, order.last_name, order.order_phone, order.order_username, order.order_date.strftime('%Y-%m-%d'), order.order_address]
            data.append(row)

        table = Table(data)

        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),

            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 14),

            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND',(0,1),(-1,-1),colors.beige),
            ('GRID',(0,0),(-1,-1),1,colors.black)
        ])
        table.setStyle(style)

        elements = []
        elements.append(table)

        doc.build(elements)
    else:
        return HttpResponse("Invalid format")

    return response

def create_order(request):
    user_username = request.user.username
    cart_id = request.session.get('cart_id')

    if request.method == 'POST':
        print(request.POST)
        form = OrderForm(request.POST)
        if form.is_valid():
            # Create and save the order
            order = Order(
                first_name=form.cleaned_data['first_name'],
                last_name=form.cleaned_data['last_name'],
                order_username=user_username,
                order_address=form.cleaned_data['order_address'],
                user=request.user
            )
            order.save()  # Important: save first

            # Then create OrderItems
            try:
                cart = Cart.objects.get(cart_id=cart_id)
                cart_items = CartItem.objects.filter(cart=cart)
                for item in cart_items:
                    OrderItem.objects.create(
                        order=order,
                        product=item.product,
                        product_name=item.product_name,
                        quantity=item.quantity,
                    )
                clear_cart(request)  # Optionally clear cart after ordering
            except Cart.DoesNotExist:
                pass  # You can add an error message here

            return redirect('order_complete')  # Redirect to a success page
        else:
            return render(request, 'orders.html', {'form': form})

    else:
        form = OrderForm()
    return render(request, 'orders.html', {'form': form})

def export_view(request):
    export_orders_to_gsheets()
    return HttpResponse("Data exported to Google Sheets")